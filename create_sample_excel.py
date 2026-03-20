#!/usr/bin/env python
"""
Create sample Excel file for Phase 7 bulk payment testing
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Payments"

# Add headers
headers = ['Student ID', 'Amount', 'Receipt Number', 'Payment Method']
ws.append(headers)

# Style header row
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Add sample data (5 students)
sample_data = [
    ['STU_BULK_001', 2500.00, 'RCP_BULK_001', 'CASH'],
    ['STU_BULK_002', 3000.00, 'RCP_BULK_002', 'BANK_TRANSFER'],
    ['STU_BULK_003', 2500.00, 'RCP_BULK_003', 'MOBILE_MONEY'],
    ['STU_BULK_004', 5000.00, 'RCP_BULK_004', 'CHEQUE'],
    ['STU_BULK_005', 2500.00, 'RCP_BULK_005', 'CASH'],
]

for row in sample_data:
    ws.append(row)

# Format columns
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 16

# Save
filename = 'sample_bulk_payments.xlsx'
wb.save(filename)
print(f'Created {filename}')
